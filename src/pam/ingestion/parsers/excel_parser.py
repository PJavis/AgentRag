"""
ExcelParser: chuyển .xlsx / .xls / .csv thành markdown hoặc CSV text.

Chiến lược:
  mode="markdown" — mỗi sheet thành markdown table, chunk như text thường.
                    Phù hợp khi cần đọc hiểu ngữ nghĩa dữ liệu.
  mode="sql"      — mỗi sheet thành CSV text với header rõ ràng.
                    StructuredReasoningPipeline sẽ tự extract schema và query.

Dependencies cần cài:
  uv add openpyxl        # cho .xlsx
  uv add xlrd            # cho .xls (legacy)
  csv là built-in Python
"""
from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Literal


def _xlsx_to_sheets(file_path: str) -> list[tuple[str, list[list[str]]]]:
    """Trả về list[(sheet_name, rows)] với rows là list[list[str]]."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel parsing. Run: uv add openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets: list[tuple[str, list[list[str]]]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [
            [str(cell.value) if cell.value is not None else "" for cell in row]
            for row in ws.iter_rows()
        ]
        # Bỏ trailing empty rows
        while rows and all(cell == "" for cell in rows[-1]):
            rows.pop()
        if rows:
            sheets.append((sheet_name, rows))
    wb.close()
    return sheets


def _xls_to_sheets(file_path: str) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(
            "xlrd is required for .xls parsing. Run: uv add xlrd"
        ) from exc

    wb = xlrd.open_workbook(file_path)
    sheets: list[tuple[str, list[list[str]]]] = []
    for sheet in wb.sheets():
        rows = [
            [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        if rows:
            sheets.append((sheet.name, rows))
    return sheets


def _csv_to_sheets(file_path: str) -> list[tuple[str, list[list[str]]]]:
    path = Path(file_path)
    with path.open(newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return [(path.stem, rows)] if rows else []


def _rows_to_markdown(rows: list[list[str]], rows_per_batch: int = 8) -> str:
    """Chuyển rows thành markdown table(s).

    Row đầu tiên là header. Nếu bảng lớn, chia thành nhiều sub-table,
    mỗi sub-table đều có header riêng để khi chunker cắt, header vẫn
    nằm cùng chunk với data rows.
    """
    if not rows:
        return ""
    header = rows[0]
    sep = ["-" * max(len(h), 3) for h in header]
    header_lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]

    data_rows = rows[1:]
    if not data_rows:
        return "\n".join(header_lines)

    # Split data into batches, each prefixed with the header
    batches: list[str] = []
    for start in range(0, len(data_rows), rows_per_batch):
        batch = data_rows[start : start + rows_per_batch]
        lines = list(header_lines)
        for row in batch:
            padded = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded[: len(header)]) + " |")
        batches.append("\n".join(lines))

    return "\n\n".join(batches)


def _rows_to_csv_text(sheet_name: str, rows: list[list[str]]) -> str:
    """Chuyển rows thành CSV text với header rõ ràng (dùng cho mode=sql)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    return f"### Sheet: {sheet_name}\n\n```csv\n{buf.getvalue()}```\n"


class ExcelParser:
    """Parse .xlsx / .xls / .csv thành text phù hợp cho chunking."""

    def parse(
        self,
        file_path: str,
        mode: Literal["markdown", "sql"] = "markdown",
    ) -> dict:
        """
        Returns:
            {
              "parsed_content": str,    # markdown hoặc csv text
              "sheets": list[str],      # tên các sheet
              "total_rows": int,        # tổng số row (không kể header)
            }
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext in (".xlsx",):
            sheets = _xlsx_to_sheets(file_path)
        elif ext in (".xls",):
            sheets = _xls_to_sheets(file_path)
        elif ext in (".csv",):
            sheets = _csv_to_sheets(file_path)
        else:
            raise ValueError(f"ExcelParser does not support extension: {ext}")

        sections: list[str] = []
        total_rows = 0

        for sheet_name, rows in sheets:
            if not rows:
                continue
            total_rows += len(rows) - 1  # trừ header

            if mode == "markdown":
                table_md = _rows_to_markdown(rows)
                sections.append(f"## Sheet: {sheet_name}\n\n{table_md}")
            else:  # sql
                sections.append(_rows_to_csv_text(sheet_name, rows))

        return {
            "parsed_content": "\n\n".join(sections),
            "sheets": [s for s, _ in sheets],
            "total_rows": total_rows,
        }
